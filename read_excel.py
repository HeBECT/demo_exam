import json
import sys
import openpyxl

def read_excel_file(filename):
    """Читает Excel файл и возвращает данные"""
    wb = openpyxl.load_workbook(filename)
    ws = wb.active
    
    # Получаем заголовки из первой строки
    headers = []
    for cell in ws[1]:
        headers.append(cell.value)
    
    # Получаем данные
    data = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if any(row):  # Пропускаем пустые строки
            row_data = {}
            for i, value in enumerate(row):
                if i < len(headers) and headers[i]:
                    row_data[headers[i]] = value
            data.append(row_data)
    
    return {"headers": headers, "data": data}

# Список файлов для чтения
files = [
    "Таблицы/Заказ покупателя.xlsx",
    "Таблицы/Производство.xlsx",
    "Таблицы/Расчет стоимости продукции.xlsx",
    "Таблицы/Спецификация.xlsx",
    "Таблицы/Цены.xlsx"
]

results = {}

for file in files:
    try:
        print(f"Чтение файла: {file}")
        data = read_excel_file(file)
        results[file] = data
        print(f"  Найдено строк: {len(data['data'])}")
        print(f"  Колонки: {data['headers']}")
        print()
    except Exception as e:
        print(f"Ошибка при чтении {file}: {e}")
        print()

# Сохраняем результаты в JSON
with open("excel_data.json", "w", encoding="utf-8") as f:
    json.dump(results, f, ensure_ascii=False, indent=2, default=str)

print("Данные сохранены в excel_data.json")
